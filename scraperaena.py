import os
import re
import json
import requests
import xlrd
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime

BASE_URL = "https://www.aena.es"
STATS_URL = "https://www.aena.es/es/estadisticas/informes-mensuales.html"
DATA_DIR = "data"
XLS_DIR = os.path.join(DATA_DIR, "xls")
JSON_PATH = os.path.join(DATA_DIR, "data.json")

def ensure_dirs():
    os.makedirs(XLS_DIR, exist_ok=True)

def _extract_arrivals(values) -> int | None:
    """
    Estrae il numero degli arrivi (Llegadas).
    Nei file Excel Aena, i numeri grandi per riga sono tipicamente in quest'ordine: 
    1. Pasajeros Totales, 2. Llegadas, 3. Salidas.
    Saltando il primo numero grande trovato, otteniamo automaticamente gli Arrivi.
    """
    first_found = False
    for v in values:
        if isinstance(v, (int, float)) and v > 10000:
            if not first_found:
                first_found = True
                continue # Salta il Totale
            return int(v) # Ritorna gli Arrivi (Llegadas)
    return None

def parse_with_xlrd(file_path: str) -> int | None:
    try:
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            # SALVATAGGIO CRITICO: Ignora i fogli "Acumulado" per evitare di prendere i dati annuali sommati
            if "ACUM" in sheet.name.upper():
                continue
                
            for row_idx in range(sheet.nrows):
                row_values = sheet.row_values(row_idx)
                if any(isinstance(cell, str) and ("PALMA DE MALLORCA" in cell.upper() or cell.upper() == "PALMA") for cell in row_values):
                    return _extract_arrivals(row_values)
    except Exception as e:
        print(f"Errore xlrd su {file_path}: {e}")
    return None

def parse_with_openpyxl(file_path: str) -> int | None:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheetname in wb.sheetnames:
            # SALVATAGGIO CRITICO: Ignora i fogli "Acumulado" per evitare di prendere i dati annuali sommati
            if "ACUM" in sheetname.upper():
                continue
                
            sheet = wb[sheetname]
            for row in sheet.iter_rows(values_only=True):
                if row and any(isinstance(cell, str) and ("PALMA DE MALLORCA" in cell.upper() or str(cell).upper() == "PALMA") for cell in row):
                    return _extract_arrivals(row)
    except Exception as e:
        print(f"Errore openpyxl su {file_path}: {e}")
    return None

def process_file(file_path: str) -> int | None:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        return parse_with_xlrd(file_path)
    elif ext == '.xlsx':
        return parse_with_openpyxl(file_path)
    return None

def scrape_and_update():
    ensure_dirs()
    
    print("Scaricando la pagina delle statistiche...")
    response = requests.get(STATS_URL)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')

    links = soup.find_all('a', href=True)
    xls_links = []
    
    print(f"DEBUG: Trovati in totale {len(links)} link generici nella pagina.")
    
    # Trova tutti i link a file Excel verificando se contengono .xls o .xlsx
    for a in links:
        href = a['href']
        if '.xls' in href.lower() or '.xlsx' in href.lower():
            full_url = urljoin(BASE_URL, href)
            xls_links.append(full_url)

    print(f"Trovati {len(xls_links)} file Excel. Estrazione dei mesi e anni...")
    
    # Alert di debug se nessun link Excel corrisponde ai criteri
    if len(xls_links) == 0:
        print("DEBUG ALERT: Nessun link Excel trovato! Mostro i primi 10 link della pagina per controllo:")
        for a in links[:10]:
            print(f" -> Testo: {a.text.strip()} | Href: {a['href']}")

    # Carichiamo i dati esistenti se ci sono
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {}

    # Regex flessibile per trovare pattern tipo ANNO_MESE o MESE_ANNO separati da underscore o trattino
    pattern_generico = re.compile(r'(\d{4})[-_](\d{2})|(\d{2})[-_](\d{4})')
    
    for url in xls_links:
        print(f"DEBUG Analisi URL: {url}")
        
        match = pattern_generico.search(url)
        if match:
            # Estrazione sicura di anno e mese
            if match.group(1): # Caso in cui viene prima l'anno (es. 2026_05)
                year = match.group(1)
                month = match.group(2)
            else: # Caso in cui viene prima il mese (es. 05_2026)
                month = match.group(3)
                year = match.group(4)
            
            # Ricaviamo l'estensione per salvare il file in modo corretto per il parser
            ext = '.xlsx' if '.xlsx' in url.lower() else '.xls'
            
            # Se l'anno non è nel dict, lo inizializziamo con 12 zeri
            if year not in data:
                data[year] = [0] * 12
                
            month_idx = int(month) - 1
            
            # Se abbiamo già un dato > 0 per quel mese/anno, saltiamo
            if data[year][month_idx] > 0:
                print(f"Dati per {year}-{month} già presenti, salto.")
                continue
                
            filename = f"{year}_{month}{ext}"
            file_path = os.path.join(XLS_DIR, filename)
            
            # Scarica il file se non esiste
            if not os.path.exists(file_path):
                print(f"Scaricando {filename} da {url}...")
                try:
                    r = requests.get(url)
                    r.raise_for_status()
                    with open(file_path, 'wb') as f:
                        f.write(r.content)
                except Exception as e:
                    print(f"Errore durante il download di {url}: {e}")
                    continue

            # Estrai gli arrivi dal file scaricato
            print(f"Estrazione dati da {filename}...")
            arrivals = process_file(file_path)
            
            if arrivals:
                data[year][month_idx] = arrivals
                print(f"Trovato: {year}-{month} -> {arrivals} arrivi")
            else:
                print(f"Nessun dato trovato per Palma in {filename}")
        else:
            print(f" -> URL ignorato (non corrisponde al pattern della data)")

    # Salva il JSON aggiornato
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    print("Dati salvati in data.json")

if __name__ == "__main__":
    scrape_and_update()
